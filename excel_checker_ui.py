import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
import threading
import time

def check_excel_file_advanced(file_path):
    try:
        # Use read_only mode for faster loading
        wb = load_workbook(file_path, data_only=True, read_only=True)
        filename = os.path.basename(file_path)
        results = []
        
        # Check required sheets first
        required_sheets = {'表紙', 'テスト項目'}
        missing_sheets = required_sheets - set(wb.sheetnames)
        if missing_sheets:
            return f"[ERROR] {filename}: \nMissing sheet(s): {', '.join(missing_sheets)}"

        ws = wb['表紙']
        
        # Step 1: Check cell P24 (確認) is not empty
        p24_value = ws['P24'].value
        if p24_value is None or str(p24_value).strip() == "":
            wb.close()  # Close workbook explicitly
            return f"[ERROR] {filename}: \nMissing \"Confirm by\"."

        results.append(f" Confirm by: {str(p24_value).strip()}")

        # Step 2: Check Status OK for each Test case in 'テスト項目'
        ws2 = wb['テスト項目']
        error_rows = []
        checked_rows = []
        status_addr = ""
        status_col = 0
        
        for cell in ws2[3]:
            if cell.value == "確認":
                status_addr = cell.coordinate[:2]
                status_col = cell.column
                break

        max_rows_to_check = 1000
        consecutive_empty_limit = 10
        consecutive_empty = 0
        
        for row in range(5, max_rows_to_check + 1):
            if consecutive_empty >= consecutive_empty_limit:
                break
                
            b_cell = ws2.cell(row=row, column=2)  # Column B
            b_value = b_cell.value
            
            if b_value is not None and str(b_value).strip():
                consecutive_empty = 0
                checked_rows.append(row)

                status_cell = ws2.cell(row=row, column=status_col)  # Column status
                status_value = status_cell.value
                if status_value is None or str(status_value).strip().upper() != "OK":
                    error_rows.append(
                        f"Row {row} (B{row}='{str(b_value).strip()}', "
                        f"{status_addr}{row}='{str(status_value or '').strip()}')"
                    )
            else:
                consecutive_empty += 1
        
        # Compile results
        if not checked_rows:
            results.append(" Did not find any Test case in sheet テスト項目")
        else:
            results.append(f" Checked {len(checked_rows)} Test case(s):")
            
            if error_rows:
                results.append(f" {len(error_rows)} rows failed validation:")
                for error_row in error_rows[:5]:  # Show first 5 errors
                    results.append(f"  - {error_row}")
                if len(error_rows) > 5:
                    results.append(f"  ... and {len(error_rows) - 5} more errors")
                wb.close()  # Close workbook explicitly
                return f"[ERROR] {filename}:\n" + "\n".join(results)
            else:
                results.append(" All 確認 rows contain value 'OK'")
        
        wb.close()  # Close workbook explicitly
        return f"[SUCCESS] {filename}:\n" + "\n".join(results)
        
    except Exception as e:
        return f"[ERROR] {os.path.basename(file_path)}: {str(e)}"

def find_excel_files_recursive(folder_path):
    """Recursively find all Excel files in folder and subfolders"""
    excel_files = []
    excel_extensions = ('.xlsx', '.xlsm', '.xls')
    
    try:
        for root_dir, _, files in os.walk(folder_path):
            for file in files:
                if file.lower().endswith(excel_extensions):
                    full_path = os.path.join(root_dir, file)
                    relative_path = os.path.relpath(full_path, folder_path)
                    excel_files.append({
                        'full_path': full_path,
                        'relative_path': relative_path,
                        'filename': file
                    })
    except PermissionError as e:
        print(f"Permission denied accessing: {e}")
    except Exception as e:
        print(f"Error scanning folders: {e}")
    
    return excel_files

def browse_folder():
    folder = filedialog.askdirectory()
    folder_path_var.set(folder)

def update_result_text(result):
    """Thread-safe way to update the result text with colored formatting"""
    def _update():
        if "[SUCCESS]" in result:
            # Insert success message in green
            result_text.insert(tk.END, result + "\n\n", "success")
        else:
            # Insert error message in red
            result_text.insert(tk.END, result + "\n\n", "error")
        result_text.see(tk.END)  # Auto-scroll to bottom
        result_text.update()
    
    root.after(0, _update)

def update_progress(current, total, current_file=""):
    """Thread-safe way to update progress bar and status"""
    def _update():
        progress_var.set((current / total) * 100)
        status_label.config(text=f"Processing: {current_file} ({current}/{total})")
        root.update_idletasks()
    
    root.after(0, _update)

def reset_ui():
    """Reset UI to initial state"""
    progress_var.set(0)
    status_label.config(text="Ready")
    check_btn.config(state="normal", text="Check Excel Files")
    browse_btn.config(state="normal")

def check_files_thread():
    """Run the checking process in a separate thread"""
    folder_path = folder_path_var.get()

    if not os.path.isdir(folder_path):
        root.after(0, lambda: messagebox.showerror("Error", "Please select a valid folder."))
        root.after(0, reset_ui)
        return

    # Show scanning message
    def show_scanning():
        result_text.delete(1.0, tk.END)
        result_text.insert(tk.END, "Scanning folder and subfolders for Excel files...\n\n", "info")
        status_label.config(text="Scanning folders...")
        root.update_idletasks()  # Force UI update
    
    root.after(0, show_scanning)

    # Find all Excel files recursively
    excel_files_info = find_excel_files_recursive(folder_path)
    
    if not excel_files_info:
        root.after(0, lambda: messagebox.showinfo("No Files", "No Excel files found in the folder or its subfolders."))
        root.after(0, reset_ui)
        return

    total_files = len(excel_files_info)
    
    # Clear results and show initial message with folder summary
    def clear_and_start():
        result_text.delete(1.0, tk.END)
        result_text.insert(tk.END, f"Found {total_files} Excel file(s) in folder and subfolders.\n", "info")
        result_text.insert(tk.END, f"Starting validation process...\n\n", "info")
        progress_var.set(0)
        root.update_idletasks()  # Force UI update
    
    root.after(0, clear_and_start)

    # Process each file
    success_count = 0
    error_count = 0
    
    for i, file_info in enumerate(excel_files_info, 1):
        full_path = file_info['full_path']
        relative_path = file_info['relative_path']
        filename = file_info['filename']
        
        # Update progress before processing
        update_progress(i-1, total_files, relative_path)
        
        # Add "Processing..." message with relative path
        def add_processing_msg(rel_path=relative_path):
            result_text.insert(tk.END, f"Processing: {rel_path}...\n", "processing")
            result_text.see(tk.END)
            root.update_idletasks()  # Force UI update
        
        root.after(0, add_processing_msg)
        
        # Check the file
        result = check_excel_file_advanced(full_path)
        
        # Count results
        if result.startswith("[SUCCESS]"):
            success_count += 1
        else:
            error_count += 1
        
        # Modify result to show relative path
        if result.startswith("[SUCCESS]") or result.startswith("[ERROR]"):
            result = result.replace(f"[SUCCESS] {filename}:", f"[SUCCESS] {relative_path}:")
            result = result.replace(f"[ERROR] {filename}:", f"[ERROR] {relative_path}:")
        
        # Update with result
        def update_final_result(res=result, rel_path=relative_path):
            lines = result_text.get(1.0, tk.END).strip().split('\n')
            if lines and f"Processing: {rel_path}..." in lines[-1]:
                result_text.delete(f"end-2l", tk.END)
            
            if "[SUCCESS]" in res:
                result_text.insert(tk.END, res + "\n\n", "success")
            else:
                result_text.insert(tk.END, res + "\n\n", "error")
            result_text.see(tk.END)
            root.update_idletasks()  # Force UI update
        
        root.after(0, update_final_result)
        
        # Small delay to make the progress visible
        time.sleep(0.05)

    # Complete the progress
    update_progress(total_files, total_files, "Complete!")
    
    # Final message
    def show_completion():
        result_text.insert(
            tk.END, 
            f"\nCompleted checking {total_files} files ({success_count} successful, {error_count} errors)!", 
            "success" if error_count == 0 else "error"
        )
        result_text.see(tk.END)
        status_label.config(text=f"Completed - {success_count} passed, {error_count} failed")
        check_btn.config(state="normal", text="Check Excel Files")
        browse_btn.config(state="normal")
        root.update_idletasks()  # Force UI update
    
    root.after(0, show_completion)

def start_check():
    """Start the checking process"""
    # Disable buttons during processing
    check_btn.config(state="disabled", text="Checking...")
    browse_btn.config(state="disabled")
    
    # Start the checking thread
    thread = threading.Thread(target=check_files_thread, daemon=True)
    thread.start()

def center_window(root, width, height):
    root.update_idletasks()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    root.geometry(f"{width}x{height}+{x}+{y}")

# --- STYLED GUI SETUP ---
root = tk.Tk()
root.title("Excel Content Checker")
center_window(root, 1240, 800)
root.configure(bg="#f0f0f0")

# Configure style
style = ttk.Style()
style.theme_use('clam')

# Custom colors
PRIMARY_COLOR = "#2c3e50"
SECONDARY_COLOR = "#3498db"
SUCCESS_COLOR = "#27ae60"
ACCENT_COLOR = "#e74c3c"
BG_COLOR = "#ecf0f1"
TEXT_COLOR = "#2c3e50"

# Configure custom styles
style.configure("Title.TLabel", font=("Segoe UI", 16, "bold"), foreground=PRIMARY_COLOR, background=BG_COLOR)
style.configure("Header.TLabel", font=("Segoe UI", 10, "bold"), foreground=PRIMARY_COLOR, background=BG_COLOR)
style.configure("Custom.TEntry", fieldbackground="white", borderwidth=1, relief="solid")
style.configure("Primary.TButton", font=("Segoe UI", 10, "bold"))

# Configure root background
root.configure(bg=BG_COLOR)

# Create main container with padding
main_frame = ttk.Frame(root, padding="20")
main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

# Configure grid weights
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)
main_frame.columnconfigure(1, weight=1)

# Title
title_label = ttk.Label(main_frame, text="Excel Content Checker", style="Title.TLabel")
title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))

# Create input section frame
input_frame = ttk.LabelFrame(main_frame, text="Configuration", padding="15")
input_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 20))
input_frame.columnconfigure(1, weight=1)

# Variables
folder_path_var = tk.StringVar()
progress_var = tk.DoubleVar()

# Folder selection
ttk.Label(input_frame, text="Folder:", style="Header.TLabel").grid(row=1, column=0, sticky="w", pady=5)
folder_entry = ttk.Entry(input_frame, textvariable=folder_path_var, width=50, style="Custom.TEntry")
folder_entry.grid(row=1, column=1, padx=(10, 5), pady=5, sticky=(tk.W, tk.E))
browse_btn = ttk.Button(input_frame, text="Browse", command=browse_folder, style="Primary.TButton")
browse_btn.grid(row=1, column=2, padx=(5, 0), pady=5)

# Check button
check_btn = ttk.Button(main_frame, text="Check Excel Files", command=start_check, style="Primary.TButton")
check_btn.grid(row=2, column=1, pady=10)

# Progress bar section
progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="15")
progress_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
progress_frame.columnconfigure(0, weight=1)

progress_bar = ttk.Progressbar(
    progress_frame, 
    variable=progress_var, 
    maximum=100, 
    mode='determinate',
    length=400
)
progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)

progress_label = ttk.Label(progress_frame, text="0%")
progress_label.grid(row=0, column=1, padx=(10, 0))

# Update progress label
def update_progress_label():
    progress_label.config(text=f"{progress_var.get():.0f}%")
    root.after(100, update_progress_label)

update_progress_label()

# Results section
results_frame = ttk.LabelFrame(main_frame, text="Results", padding="15")
results_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))
results_frame.columnconfigure(0, weight=1)
results_frame.rowconfigure(0, weight=1)

# Results text with scrollbar
text_frame = tk.Frame(results_frame, bg=BG_COLOR)
text_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
text_frame.columnconfigure(0, weight=1)
text_frame.rowconfigure(0, weight=1)

result_text = tk.Text(
    text_frame, 
    height=12, 
    width=70,
    font=("Consolas", 10),
    bg="white",
    fg=TEXT_COLOR,
    relief="flat",
    borderwidth=1,
    wrap=tk.WORD,
    padx=10,
    pady=10
)
result_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

# Configure text tags for colored text
result_text.tag_configure("error", foreground="#e74c3c", font=("Consolas", 10, "normal"))        # Red for errors
result_text.tag_configure("success", foreground="#27ae60", font=("Consolas", 10, "normal"))      # Green for success
result_text.tag_configure("processing", foreground="#3498db", font=("Consolas", 10, "italic"))  # Blue for processing
result_text.tag_configure("info", foreground="#34495e", font=("Consolas", 10, "normal"))        # Dark gray for info

# Scrollbar for results
scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=result_text.yview)
scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
result_text.configure(yscrollcommand=scrollbar.set)

# Configure main frame grid weights
main_frame.rowconfigure(4, weight=1)

# Status bar
status_frame = tk.Frame(root, bg=PRIMARY_COLOR, height=25)
status_frame.grid(row=1, column=0, sticky=(tk.W, tk.E))
status_label = tk.Label(status_frame, text="Ready", bg=PRIMARY_COLOR, fg="white", font=("Segoe UI", 9))
status_label.pack(side=tk.LEFT, padx=10, pady=2)

root.mainloop()