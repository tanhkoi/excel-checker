import sys
import os
import time
from openpyxl import load_workbook
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QPushButton,
    QVBoxLayout,
    QFileDialog,
    QLabel,
    QHBoxLayout,
    QProgressBar,
    QMessageBox,
    QLineEdit,
    QTableWidget,
    QTableWidgetItem,
    QCheckBox,
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QColor
import subprocess
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
from datetime import datetime
from pathlib import Path
import logging


# Configure logging for debugging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def load_config(config_path="config.json"):
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


CONFIG = load_config()

# --- Constants ---
CATEGORY_PREFIX_MAP = CONFIG["category_prefix_map"]
INVALID_SHEETS = set(CONFIG["invalid_sheets"])  # Use set for O(1) lookup
REQUIRED_SHEETS = set(CONFIG["required_sheets"])  # Use set for O(1) lookup
EXCEL_EXTENSIONS = tuple(CONFIG["excel_extensions"])
INVALID_CHARS = set(CONFIG["invalid_chars"])  # Use set for faster lookup
INVALID_TEXT = tuple(CONFIG["invalid_text"])  # Keep as tuple for 'in' check


# --- Optimized Helper Functions ---
def check_invalid_text(wb):
    """Check for invalid text patterns - early exit on first match"""
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            # Use row iterator for better memory efficiency
            for row in ws.iter_rows(
                values_only=True#, max_row=10000
            ):  # Limit rows for performance
                for cell in row:
                    if isinstance(cell, str) and any(
                        text in cell for text in INVALID_TEXT
                    ):
                        return f"Contains invalid text in sheet '{sheet_name}'"
        except Exception as e:
            logger.warning(f"Error checking invalid text in sheet {sheet_name}: {e}")
            continue
    return None


def column_letter(col_idx):
    """Optimized column letter conversion with caching"""
    if not hasattr(column_letter, "cache"):
        column_letter.cache = {}

    if col_idx in column_letter.cache:
        return column_letter.cache[col_idx]

    letters = []
    original_idx = col_idx
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letters.append(chr(65 + remainder))

    result = "".join(reversed(letters))
    column_letter.cache[original_idx] = result
    return result


def check_contains_vietnamese_characters(wb):
    """Optimized Vietnamese character check with early exit and row limits"""
    vietnamese_chars = INVALID_CHARS

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            # Limit to first 1000 rows for performance
            for row_idx, row in enumerate(
                ws.iter_rows(values_only=True), start=1#, max_row=1000
            ):
                for col_idx, cell in enumerate(row, start=1):
                    if isinstance(cell, str) and vietnamese_chars.intersection(
                        set(cell)
                    ):
                        col_letter = column_letter(col_idx)
                        cell_preview = cell[:50] + "..." if len(cell) > 50 else cell
                        return (
                            f"Contains Vietnamese characters at {sheet_name}!{col_letter}{row_idx} "
                            f"(value: '{cell_preview}')"
                        )
        except Exception as e:
            logger.warning(
                f"Error checking Vietnamese chars in sheet {sheet_name}: {e}"
            )
            continue
    return None


def check_valid_filename(file_path):
    """Optimized filename check using pathlib"""
    path_obj = Path(file_path)
    filename = path_obj.name
    path_parts = path_obj.parts

    for folder_name, expected_prefix in CATEGORY_PREFIX_MAP.items():
        if folder_name in path_parts:
            if not filename.startswith(expected_prefix):
                return f"Invalid filename for '{folder_name}'"
            break
    return None


def check_invalid_sheet(wb):
    """Use set intersection for faster sheet validation"""
    invalid_found = INVALID_SHEETS.intersection(wb.sheetnames)
    if invalid_found:
        return f"Contains invalid sheet: {invalid_found.pop()}"
    return None


def check_required_sheets(wb):
    """Use set operations for faster sheet validation"""
    missing_sheets = REQUIRED_SHEETS - set(wb.sheetnames)
    if missing_sheets:
        return f"Missing required sheet: {missing_sheets.pop()}"
    return None


def check_confirm_by(wb):
    """Optimized confirm check with error handling"""
    if "表紙" not in wb.sheetnames:
        return None

    try:
        ws = wb["表紙"]
        cell_value = ws["P24"].value
        return (
            "Missing Confirm"
            if cell_value is None or not str(cell_value).strip()
            else None
        )
    except Exception as e:
        logger.warning(f"Error checking confirm cell: {e}")
        return "Error reading confirm cell"


def find_column_indexes(ws, headers=("確認", "参考"), header_row=3):
    """Optimized column index finding"""
    try:
        return {
            cell.value: cell.column for cell in ws[header_row] if cell.value in headers
        }
    except Exception:
        return {}


def check_status_in_test_items(wb, max_rows=500, empty_limit=10):  # Reduced max_rows
    """Optimized test item status check with better early termination"""
    if "テスト項目" not in wb.sheetnames:
        return None

    try:
        ws = wb["テスト項目"]
        col_indexes = find_column_indexes(ws)
        if "確認" not in col_indexes:
            return "Column '確認' not found"

        status_col = col_indexes["確認"]
        error_rows = []
        consecutive_empty = 0

        # Use iter_rows for better performance
        for row in ws.iter_rows(min_row=5, max_row=max_rows, values_only=False):
            if consecutive_empty >= empty_limit:
                break

            b_cell = row[1]  # Column B (index 1)
            if b_cell.value and str(b_cell.value).strip():
                consecutive_empty = 0
                # Get status cell by column index
                status_cell = ws.cell(row=b_cell.row, column=status_col)
                status_value = status_cell.value
                if status_value is None or str(status_value).strip().upper() != "OK":
                    error_rows.append(str(b_cell.value).strip())
                    # Limit error collection for performance
                    if len(error_rows) > 100:  # Stop collecting after 100 errors
                        error_rows.append("... (more errors)")
                        break
            else:
                consecutive_empty += 1

        return (
            f"{len(error_rows)} TC(s) != 'OK': "
            + "; ".join(error_rows[:10])  # Limit display
            if error_rows
            else None
        )
    except Exception as e:
        logger.warning(f"Error checking test items: {e}")
        return f"Error checking test items: {e}"


# --- Optimized Main Function ---
def check_excel_file_advanced(file_path, options):
    """Optimized Excel file checking with better error handling and performance"""
    try:
        error_messages = []

        # Use optimized openpyxl parameters for better performance
        wb = load_workbook(
            file_path,
            data_only=True,
            read_only=True,
            keep_links=False,  # Disable external links for performance
        )

        # Quick filename check first (no file I/O)
        if options.get("check_filename_prefix", True):
            if err := check_valid_filename(file_path):
                error_messages.append(err)

        # Sheet existence checks (fast)
        if err := check_required_sheets(wb):
            error_messages.append(err)

        if options.get("check_invalid_sheets", True):
            if err := check_invalid_sheet(wb):
                error_messages.append(err)

        # Cell content checks (slower, do them last)
        if options.get("check_confirm_cell", True):
            if err := check_confirm_by(wb):
                error_messages.append(err)

        if options.get("check_testcase_status", True):
            if err := check_status_in_test_items(wb):
                error_messages.append(err)

        # Most expensive checks last
        if options.get("check_contains_vietnamese_characters", True):
            if err := check_contains_vietnamese_characters(wb):
                error_messages.append(err)

        if options.get("check_invalid_text", True):
            if err := check_invalid_text(wb):
                error_messages.append(err)

        wb.close()
        return ("ERROR", ", ".join(error_messages)) if error_messages else ("OK", "")

    except Exception as e:
        logger.error(f"Error processing {file_path}: {e}")
        return "ERROR", str(e)


def find_excel_files_recursive(folder_path):
    """Optimized file finding using pathlib and generator"""
    folder = Path(folder_path)
    excel_files = []

    try:
        # Use pathlib.rglob for better performance
        for file_path in folder.rglob("*"):
            if file_path.is_file() and file_path.suffix.lower() in [
                ext.lower() for ext in EXCEL_EXTENSIONS
            ]:
                excel_files.append(str(file_path))
    except Exception as e:
        logger.error(f"Error scanning folder {folder_path}: {e}")

    return excel_files


# ----------------- Optimized Worker Thread -----------------
class ExcelCheckWorker(QThread):
    progress_changed = pyqtSignal(int)
    file_result = pyqtSignal(str, str, str, str)
    finished_signal = pyqtSignal()

    def __init__(self, folder_path, options, max_workers=None):
        super().__init__()
        self.folder_path = folder_path
        self.options = options
        # Optimize worker count based on CPU cores
        if max_workers is None:
            max_workers = min(8, (os.cpu_count() or 1) + 4)
        self.max_workers = max_workers
        self._is_running = True

    def run(self):
        start_time = time.time()
        files = find_excel_files_recursive(self.folder_path)
        total = len(files)

        if not files:
            self.file_result.emit(self.folder_path, "", "INFO", "No Excel files found.")
            self.finished_signal.emit()
            return

        logger.info(f"Processing {total} files with {self.max_workers} workers")
        processed = 0

        # Use optimized ThreadPoolExecutor settings
        with ThreadPoolExecutor(
            max_workers=self.max_workers, thread_name_prefix="ExcelChecker"
        ) as executor:
            # Submit all tasks
            futures = {
                executor.submit(check_excel_file_advanced, file, self.options): file
                for file in files
            }

            # Process results as they complete
            for future in as_completed(futures):
                if not self._is_running:
                    break

                file_path = futures[future]
                try:
                    relative_path = os.path.relpath(file_path, self.folder_path)
                    status, error_msg = future.result(timeout=30)  # Add timeout

                    self.file_result.emit(
                        self.folder_path, relative_path, status, error_msg
                    )
                except Exception as e:
                    logger.error(f"Error processing {file_path}: {e}")
                    relative_path = os.path.relpath(file_path, self.folder_path)
                    self.file_result.emit(
                        self.folder_path, relative_path, "ERROR", str(e)
                    )

                processed += 1
                progress = int((processed / total) * 100)
                self.progress_changed.emit(progress)

        elapsed = time.time() - start_time
        logger.info(f"Processing completed in {elapsed:.2f} seconds")
        self.finished_signal.emit()

    def stop(self):
        self._is_running = False


# ----------------- Optimized PyQt Main Window -----------------
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Checker - Optimized")
        self.setGeometry(100, 100, 1200, 700)  # Larger window
        self.worker = None

        # Performance tracking
        self.start_time = None

        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout()

        # Input section
        input_layout = QHBoxLayout()
        self.folder_input = QLineEdit()
        self.folder_input.setPlaceholderText("Paste or type folder path here...")
        self.folder_input.textChanged.connect(self.on_folder_input_change)

        self.btn_select = QPushButton("Browse")
        self.btn_select.clicked.connect(self.select_folder)

        input_layout.addWidget(self.folder_input)
        input_layout.addWidget(self.btn_select)

        # Options section with better layout
        option_layout = QVBoxLayout()
        self.confirm_cell_cb = QCheckBox("1. Check confirm")
        self.testcase_status_cb = QCheckBox("2. Check test case status != 'OK'")
        self.filename_check_cb = QCheckBox("3. Check filename prefix")
        self.sheet_check_cb = QCheckBox("4. Check invalid/missing sheets")
        self.check_contains_vietnamese_characters_cb = QCheckBox(
            "5. Check Vietnamese characters (performance impact)"
        )
        self.check_invalid_text_cb = QCheckBox("6. Check invalid text patterns")

        # Set defaults with performance considerations
        self.confirm_cell_cb.setChecked(True)
        self.testcase_status_cb.setChecked(True)
        self.filename_check_cb.setChecked(True)
        self.sheet_check_cb.setChecked(True)
        self.check_contains_vietnamese_characters_cb.setChecked(False)  # Expensive
        self.check_invalid_text_cb.setChecked(False)  # Expensive

        for cb in [
            self.confirm_cell_cb,
            self.testcase_status_cb,
            self.filename_check_cb,
            self.sheet_check_cb,
            self.check_contains_vietnamese_characters_cb,
            self.check_invalid_text_cb,
        ]:
            option_layout.addWidget(cb)

        # Button section
        button_layout = QHBoxLayout()
        self.btn_execute = QPushButton("Execute")
        self.btn_execute.setEnabled(False)
        self.btn_execute.clicked.connect(self.start_execution)

        self.btn_stop = QPushButton("Stop")
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self.stop_execution)

        self.btn_export = QPushButton("Export Results")
        self.btn_export.setEnabled(False)
        self.btn_export.clicked.connect(self.export_results)

        button_layout.addWidget(self.btn_execute)
        button_layout.addWidget(self.btn_stop)
        button_layout.addWidget(self.btn_export)

        # Table widget with optimizations
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(
            ["Prefix Path", "Relative Path", "Status", "Errors"]
        )
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.itemDoubleClicked.connect(self.open_selected_file)
        self.table.setSortingEnabled(True)

        # Optimize table performance
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setAlignment(Qt.AlignCenter)
        self.progress_bar.setValue(0)

        # Status label with performance info
        config_info_label = QLabel(
            "Note: Case 3, 4, 5, and 6 are configurable.\n"
            "Options 5 and 6 have significant performance impact on large files.\n"
            "Configuration can be modified in 'config.json'."
        )
        config_info_label.setStyleSheet("color: gray; font-size: 11px;")
        config_info_label.setWordWrap(True)

        self.status_label = QLabel("Ready")

        # Assemble main layout
        main_layout.addLayout(input_layout)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(config_info_label)
        main_layout.addLayout(option_layout)
        main_layout.addWidget(self.table)
        main_layout.addWidget(self.progress_bar)
        main_layout.addWidget(self.status_label)

        self.setLayout(main_layout)

    def on_folder_input_change(self, text):
        path = text.strip()
        self.btn_execute.setEnabled(os.path.isdir(path))

    def select_folder(self):
        current_path = self.folder_input.text().strip()
        start_dir = (
            current_path if os.path.isdir(current_path) else os.path.expanduser("~")
        )

        if folder := QFileDialog.getExistingDirectory(self, "Select Folder", start_dir):
            self.folder_input.setText(folder)

    def start_execution(self):
        self.btn_export.setEnabled(False)
        folder_path = self.folder_input.text().strip()
        if not os.path.isdir(folder_path):
            QMessageBox.warning(
                self, "Invalid Folder", "Please provide a valid folder path."
            )
            return

        # Performance tracking
        self.start_time = time.time()

        self.progress_bar.setValue(0)
        self.table.setRowCount(0)
        self.btn_execute.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.status_label.setText("Scanning files...")

        options = {
            "check_invalid_sheets": self.sheet_check_cb.isChecked(),
            "check_filename_prefix": self.filename_check_cb.isChecked(),
            "check_confirm_cell": self.confirm_cell_cb.isChecked(),
            "check_testcase_status": self.testcase_status_cb.isChecked(),
            "check_contains_vietnamese_characters": self.check_contains_vietnamese_characters_cb.isChecked(),
            "check_invalid_text": self.check_invalid_text_cb.isChecked(),
        }

        # Reload config for any changes
        global CONFIG
        try:
            CONFIG = load_config()
        except Exception as e:
            logger.warning(f"Could not reload config: {e}")

        self.worker = ExcelCheckWorker(folder_path, options)
        self.worker.progress_changed.connect(self.progress_bar.setValue)
        self.worker.file_result.connect(self.add_table_row)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.start()

    def stop_execution(self):
        if self.worker:
            self.worker.stop()
            self.worker.wait(5000)  # Wait up to 5 seconds
            self.status_label.setText("Process stopped by user")
            self.btn_execute.setEnabled(True)
            self.btn_stop.setEnabled(False)

    def add_table_row(self, prefix_path, path, status, error):
        row = self.table.rowCount()
        self.table.insertRow(row)

        # Create items with optimized formatting
        items = [
            QTableWidgetItem(prefix_path),
            QTableWidgetItem(path.replace("\\", "/")),
            QTableWidgetItem(status),
            QTableWidgetItem(
                error[:500] + "..." if len(error) > 500 else error
            ),  # Truncate long errors
        ]

        # Color coding
        if status == "OK":
            items[2].setForeground(QColor("green"))
        elif status == "ERROR":
            items[2].setForeground(QColor("red"))
        elif status == "INFO":
            items[2].setForeground(QColor("blue"))

        for col, item in enumerate(items):
            self.table.setItem(row, col, item)

        # Enable export after first result
        if row == 0:
            self.btn_export.setEnabled(True)

        # Update status with current progress
        if row % 10 == 0:  # Update every 10 files to reduce UI overhead
            self.status_label.setText(f"Processing... ({row + 1} files processed)")

    def open_selected_file(self, item):
        row = item.row()
        if row >= self.table.rowCount():
            return

        path_item = self.table.item(row, 1)
        if not path_item:
            return

        path = os.path.join(self.folder_input.text(), path_item.text())

        if os.path.exists(path):
            try:
                modifiers = QApplication.keyboardModifiers()
                if modifiers == Qt.ControlModifier:
                    # Open folder
                    folder_path = os.path.dirname(path)
                    if os.name == "nt":
                        os.startfile(folder_path)
                    elif sys.platform == "darwin":
                        subprocess.call(["open", folder_path])
                    else:
                        subprocess.call(["xdg-open", folder_path])
                else:
                    # Open file
                    if os.name == "nt":
                        os.startfile(path)
                    elif sys.platform == "darwin":
                        subprocess.call(["open", path])
                    else:
                        subprocess.call(["xdg-open", path])
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Could not open: {str(e)}")
        else:
            QMessageBox.warning(self, "File Not Found", f"File not found: {path}")

    def on_finished(self):
        self.btn_execute.setEnabled(True)
        self.btn_stop.setEnabled(False)

        # Show performance info
        if self.start_time:
            elapsed = time.time() - self.start_time
            file_count = self.table.rowCount()
            rate = file_count / elapsed if elapsed > 0 else 0
            self.status_label.setText(
                f"Completed: {file_count} files in {elapsed:.1f}s ({rate:.1f} files/sec)"
            )
        else:
            self.status_label.setText("Process completed")

        QMessageBox.information(self, "Done", "Check completed.")

    def closeEvent(self, event):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.worker.wait(3000)  # Wait up to 3 seconds
        event.accept()

    def export_results(self):
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "There are no results to export.")
            return

        default_name = (
            f"Excel_Check_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Results", default_name, "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Check Results"

            # Write headers with styling
            headers = ["Prefix Path", "Relative Path", "Status", "Errors"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(
                    start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
                )

            # Write data with conditional formatting
            for row in range(self.table.rowCount()):
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    cell = ws.cell(
                        row=row + 2, column=col + 1, value=item.text() if item else ""
                    )

                    # Color code status column
                    if col == 2 and item:  # Status column
                        if item.text() == "OK":
                            cell.fill = PatternFill(
                                start_color="90EE90",
                                end_color="90EE90",
                                fill_type="solid",
                            )
                        elif item.text() == "ERROR":
                            cell.fill = PatternFill(
                                start_color="FFB6C1",
                                end_color="FFB6C1",
                                fill_type="solid",
                            )

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column[
                    : min(100, len(list(column)))
                ]:  # Limit for performance
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 100)  # Cap width
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            QMessageBox.information(
                self, "Success", f"Results exported to:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(
                self, "Export Error", f"Failed to export results:\n{str(e)}"
            )


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
