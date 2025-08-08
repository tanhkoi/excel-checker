import sys
import os
import re
import json
import zipfile
from datetime import datetime
from threading import Event
from concurrent.futures import ThreadPoolExecutor, as_completed
import xml.etree.ElementTree as ET
import subprocess

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QPushButton,
    QVBoxLayout,
    QFileDialog,
    QLabel,
    QHBoxLayout,
    QProgressBar,
    QMessageBox,
    QLineEdit,
    QTableWidget,
    QTableWidgetItem,
    QCheckBox,
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QColor


# ==================== CONFIGURATION ====================
def load_config(config_path="config.json"):
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


CONFIG = load_config()

# Constants
CATEGORY_PREFIX_MAP = CONFIG["category_prefix_map"]
INVALID_SHEETS = set(CONFIG["invalid_sheets"])
REQUIRED_SHEETS = set(CONFIG["required_sheets"])
EXCEL_EXTENSIONS = tuple(CONFIG["excel_extensions"])
INVALID_CHARS = set(CONFIG["invalid_chars"])
INVALID_TEXT = set(CONFIG["invalid_text"])


# ==================== UTILITY FUNCTIONS ====================
def col_num_to_letter(col_num):
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + col_num % 26) + result
        col_num //= 26
    return result


def find_excel_files_recursive(folder_path):
    excel_files = []
    for root_dir, _, files in os.walk(folder_path):
        for file in files:
            lower_file = file.lower()
            if lower_file.startswith("~$"):
                continue
            if lower_file.endswith(EXCEL_EXTENSIONS):
                excel_files.append(os.path.join(root_dir, file))
    return excel_files


# ==================== EXCEL FILE CHECKING ====================
def get_shared_strings(zip_ref):
    try:
        with zip_ref.open("xl/sharedStrings.xml") as f:
            tree = ET.parse(f)
            root = tree.getroot()
            ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            return [
                "".join(t.text for t in si.findall(".//a:t", ns) if t.text)
                for si in root.findall("a:si", ns)
            ]
    except KeyError:
        return []


def get_sheet_names(zip_ref):
    with zip_ref.open("xl/workbook.xml") as f:
        tree = ET.parse(f)
        root = tree.getroot()
        ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        return [sheet.attrib["name"] for sheet in root.findall(".//a:sheet", ns)]


def read_cells_from_sheet(zip_ref, sheet_filename, shared_strings):
    with zip_ref.open(sheet_filename) as f:
        tree = ET.parse(f)
        root = tree.getroot()
        ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        values = []

        for c in root.iter(
            "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c"
        ):
            cell_ref = c.attrib.get("r")
            cell_type = c.attrib.get("t")
            v = c.find("a:v", ns)
            if v is not None:
                value = v.text
                if cell_type == "s":
                    value = shared_strings[int(value)]
                values.append((cell_ref, value))
        return values


def check_confirm_by(zip_ref, shared_strings, sheet_names):
    try:
        if "表紙" not in sheet_names:
            return "Missing required sheet: '表紙'"

        sheet_idx = sheet_names.index("表紙") + 1
        sheet_file = f"xl/worksheets/sheet{sheet_idx}.xml"

        with zip_ref.open(sheet_file) as f:
            tree = ET.parse(f)
            root = tree.getroot()
            ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            cells = {}

            for c in root.iter(
                "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c"
            ):
                cell_ref = c.attrib.get("r")
                cell_type = c.attrib.get("t")
                v = c.find("a:v", ns)
                value = None
                if v is not None:
                    value = v.text
                    if cell_type == "s":
                        value = shared_strings[int(value)]
                cells[cell_ref] = value

            for cell_ref, value in cells.items():
                if value == "確認":
                    m = re.match(r"([A-Z]+)([0-9]+)", cell_ref)
                    if not m:
                        continue
                    col, row = m.group(1), int(m.group(2))
                    below_ref = f"{col}{row+1}"
                    below_value = cells.get(below_ref)
                    if below_value is None:
                        return "Missing Confirm"
                    return None

    except Exception as e:
        return f"Error in check_confirm_by: {e}"
    return None


def check_status_in_test_items(
    zip_ref, shared_strings, sheet_names, max_rows=1000, empty_limit=10
):
    try:
        if "テスト項目" not in sheet_names:
            return "Missing required sheet: 'テスト項目'"

        sheet_idx = sheet_names.index("テスト項目") + 1
        sheet_file = f"xl/worksheets/sheet{sheet_idx}.xml"

        with zip_ref.open(sheet_file) as f:
            tree = ET.parse(f)
            root = tree.getroot()
            ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            cells = {}

            for c in root.iter(
                "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c"
            ):
                cell_ref = c.attrib.get("r")
                cell_type = c.attrib.get("t")
                v = c.find("a:v", ns)
                value = None
                if v is not None:
                    value = v.text
                    if cell_type == "s":
                        value = shared_strings[int(value)]
                cells[cell_ref] = value

            # Find confirmation column
            confirm_col = None
            for header_row in [3, 4]:
                for col in range(50, 100):
                    col_letter = col_num_to_letter(col)
                    cell_ref = f"{col_letter}{header_row}"
                    cell_value = cells.get(cell_ref)
                    if cell_value == "確認":
                        confirm_col = col_letter
                        break
                if confirm_col:
                    break

            if not confirm_col:
                return "Column '確認' not found"

            # Check test case statuses
            error_rows = []
            consecutive_empty = 0
            for row in range(5, max_rows + 1):
                b_cell_ref = f"B{row}"
                b_value = cells.get(b_cell_ref)

                if b_value and str(b_value).strip():
                    consecutive_empty = 0
                    status_cell_ref = f"{confirm_col}{row}"
                    status_value = cells.get(status_cell_ref)
                    if (
                        status_value is None
                        or str(status_value).strip().upper() != "OK"
                    ):
                        error_rows.append(str(b_value).strip())
                else:
                    consecutive_empty += 1
                    if consecutive_empty >= empty_limit:
                        break

            return (
                f"{len(error_rows)} TC(s) != 'OK': " + "; ".join(error_rows)
                if error_rows
                else None
            )
    except Exception as e:
        return f"Error in check_status_in_test_items: {e}"
    return None


def check_invalid_text(cell_values, sheet_name, invalid_text_set):
    for cell_ref, value in cell_values:
        if isinstance(value, str) and any(t in value for t in invalid_text_set):
            return f"{sheet_name}: Contains invalid text: {cell_ref}->{value}"
    return None


def check_contains_vn_chars(cell_values, sheet_name, invalid_chars):
    results = []
    pattern = re.compile(f"[{''.join(re.escape(c) for c in invalid_chars)}]")

    for cell_ref, cell_value in cell_values:
        if isinstance(cell_value, str) and pattern.search(cell_value):
            results.append(f"{sheet_name}: {cell_ref}->{cell_value}")

    return " ".join(results) if results else None


def check_incorrect_textbox(zip_ref):
    try:
        with zip_ref.open("xl/drawings/drawing1.xml") as f:
            tree = ET.parse(f)
            ns = {
                "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            }
            root = tree.getroot()
            texts = []

            for txBody in root.findall(".//xdr:txBody", ns):
                for p in txBody.findall(".//a:p", ns):
                    text_parts = [t.text for t in p.findall(".//a:t", ns) if t.text]
                    if text_parts:
                        texts.append("".join(text_parts))

            for text in texts:
                if not text or "API" in text:
                    return f"Incorrect TextBox content: '{text}'"

    except KeyError:
        pass
    return None


def check_valid_filename(file_path):
    filename = os.path.basename(file_path)
    parts = os.path.normpath(file_path).split(os.sep)

    for folder_name, expected_prefix in CATEGORY_PREFIX_MAP.items():
        if folder_name in parts:
            if not filename.startswith(expected_prefix):
                return f"Invalid filename for '{folder_name}'"
            break
    return None


def check_excel_file_advanced(file_path, options, stop_event=None):
    if stop_event and stop_event.is_set():
        return "CANCELLED", "Stopped by user"

    try:
        error_messages = []

        with zipfile.ZipFile(file_path, "r") as zip_ref:
            shared_strings = get_shared_strings(zip_ref)
            sheet_names = get_sheet_names(zip_ref)
            sheet_files = [
                f
                for f in zip_ref.namelist()
                if f.startswith("xl/worksheets/sheet") and f.endswith(".xml")
            ]

            # Perform checks based on options
            if options.get("check_filename_prefix", True):
                if err := check_valid_filename(file_path):
                    error_messages.append(err)

            if options.get("check_invalid_sheets", True):
                for sheet in INVALID_SHEETS:
                    if sheet in sheet_names:
                        error_messages.append(f"Contains invalid sheet: {sheet}")

            if options.get("check_required_sheets", True):
                for sheet in REQUIRED_SHEETS:
                    if sheet not in sheet_names:
                        error_messages.append(f"Missing required sheet: {sheet}")

            # Process each sheet
            for idx, sheet_file in enumerate(sheet_files):
                if stop_event and stop_event.is_set():
                    return "CANCELLED", "Stopped by user"

                cell_values = read_cells_from_sheet(zip_ref, sheet_file, shared_strings)
                sheet_name = sheet_names[idx] if idx < len(sheet_names) else sheet_file

                if options.get("check_invalid_text", True):
                    if err := check_invalid_text(cell_values, sheet_name, INVALID_TEXT):
                        error_messages.append(err)

                if options.get("check_contains_vietnamese_characters", True):
                    if err := check_contains_vn_chars(
                        cell_values, sheet_name, INVALID_CHARS
                    ):
                        error_messages.append(err)

            # Additional checks
            if options.get("check_confirm_cell", True):
                if err := check_confirm_by(zip_ref, shared_strings, sheet_names):
                    error_messages.append(err)

            if options.get("check_testcase_status", True):
                if err := check_status_in_test_items(
                    zip_ref, shared_strings, sheet_names
                ):
                    error_messages.append(err)

            if options.get("check_incorrect_tb_content", True):
                if err := check_incorrect_textbox(zip_ref):
                    error_messages.append(err)

        return ("ERROR", ", ".join(error_messages)) if error_messages else ("OK", "")

    except Exception as e:
        return "ERROR", str(e)


# ==================== WORKER THREAD ====================
class ExcelCheckWorker(QThread):
    progress_changed = pyqtSignal(int)
    file_result = pyqtSignal(
        str, str, str, str
    )  # prefix_path, relative_path, status, error
    finished_signal = pyqtSignal()

    def __init__(self, folder_path, options, max_workers=4):
        super().__init__()
        self.folder_path = folder_path
        self.options = options
        self.max_workers = max_workers
        self._stop_event = Event()

    def run(self):
        files = find_excel_files_recursive(self.folder_path)
        total = len(files)
        if not files:
            self.file_result.emit(self.folder_path, "", "INFO", "No Excel files found.")
            self.finished_signal.emit()
            return

        processed = 0
        chunk_size = 4  # Process files in chunks for better progress reporting

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            for i in range(0, total, chunk_size):
                if self._stop_event.is_set():
                    break

                current_chunk = files[i : i + chunk_size]
                chunk_futures = {
                    executor.submit(
                        check_excel_file_advanced,
                        file,
                        self.options,
                        self._stop_event,
                    ): file
                    for file in current_chunk
                }

                for future in as_completed(chunk_futures):
                    if self._stop_event.is_set():
                        break

                    file_path = chunk_futures[future]
                    try:
                        relative_path = os.path.relpath(file_path, self.folder_path)
                        status, error_msg = future.result()
                        self.file_result.emit(
                            self.folder_path, relative_path, status, error_msg
                        )
                    except Exception as e:
                        self.file_result.emit(
                            self.folder_path,
                            os.path.relpath(file_path, self.folder_path),
                            "ERROR",
                            str(e),
                        )

                    processed += 1
                    self.progress_changed.emit(int((processed / total) * 100))

        self.finished_signal.emit()

    def stop(self):
        self._stop_event.set()


# ==================== MAIN WINDOW ====================
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Checker")
        self.setGeometry(100, 100, 1000, 600)
        self.worker = None
        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout()

        # Input section
        input_layout = QHBoxLayout()
        self.folder_input = QLineEdit()
        self.folder_input.setPlaceholderText("Paste or type folder path here...")
        self.folder_input.textChanged.connect(self.on_folder_input_change)

        self.btn_select = QPushButton("Browse")
        self.btn_select.clicked.connect(self.select_folder)

        input_layout.addWidget(self.folder_input)
        input_layout.addWidget(self.btn_select)

        button_layout = QHBoxLayout()

        self.btn_execute = QPushButton("Execute")
        self.btn_execute.setEnabled(False)
        self.btn_execute.clicked.connect(self.start_execution)

        self.btn_stop = QPushButton("Stop")
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self.stop_execution)

        self.btn_export = QPushButton("Export")
        self.btn_export.setEnabled(False)
        self.btn_export.clicked.connect(self.export_results)

        self.btn_select_all = QPushButton("Select All")
        self.btn_select_all.clicked.connect(self.select_all_options)

        self.btn_deselect_all = QPushButton("Deselect All")
        self.btn_deselect_all.clicked.connect(self.deselect_all_options)

        button_layout.addWidget(self.btn_select_all)
        button_layout.addWidget(self.btn_deselect_all)
        button_layout.addStretch()
        button_layout.addWidget(self.btn_export)
        button_layout.addWidget(self.btn_execute)
        button_layout.addWidget(self.btn_stop)

        options_group = QWidget()
        options_layout = QVBoxLayout(options_group)
        options_layout.setContentsMargins(5, 5, 5, 5)

        self.confirm_cell_cb = QCheckBox("1. Check confirm")
        self.sheet_req_check_cb = QCheckBox("2. Check required sheets")
        self.testcase_status_cb = QCheckBox("3. Check test case status")
        self.filename_check_cb = QCheckBox("4. Check filename prefix")
        self.sheet_check_cb = QCheckBox("5. Check invalid sheets")
        self.check_contains_vietnamese_characters_cb = QCheckBox(
            "6. Check Vietnamese chars"
        )
        self.check_invalid_text_cb = QCheckBox("7. Check invalid text")
        self.check_incorrect_tb_content_cb = QCheckBox("8. Check Text Box content")

        # Set default states
        self.confirm_cell_cb.setChecked(False)
        self.testcase_status_cb.setChecked(False)
        self.filename_check_cb.setChecked(False)
        self.sheet_req_check_cb.setChecked(False)
        self.sheet_check_cb.setChecked(False)
        self.check_contains_vietnamese_characters_cb.setChecked(False)
        self.check_invalid_text_cb.setChecked(False)
        self.check_incorrect_tb_content_cb.setChecked(False)

        options_layout.addWidget(self.confirm_cell_cb)
        options_layout.addWidget(self.sheet_req_check_cb)
        options_layout.addWidget(self.testcase_status_cb)
        options_layout.addWidget(self.filename_check_cb)
        options_layout.addWidget(self.sheet_check_cb)
        options_layout.addWidget(self.check_contains_vietnamese_characters_cb)
        options_layout.addWidget(self.check_invalid_text_cb)
        options_layout.addWidget(self.check_incorrect_tb_content_cb)

        # Table widget
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(
            ["Prefix Path", "Relative Path", "Status", "Errors"]
        )
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.itemDoubleClicked.connect(self.open_selected_file)
        self.table.setSortingEnabled(True)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setAlignment(Qt.AlignCenter)
        self.progress_bar.setValue(0)

        # Status label
        config_info_label = QLabel(
            "Note: Case 2, 4, 5, 6 and 7 are configurable in 'config.json'"
        )
        config_info_label.setStyleSheet("color: gray; font-size: 11px;")
        config_info_label.setWordWrap(True)

        self.status_label = QLabel("Ready")

        # Assemble main layout
        main_layout.addLayout(input_layout)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(options_group)
        main_layout.addWidget(self.table)
        main_layout.addWidget(self.progress_bar)
        main_layout.addWidget(config_info_label)
        main_layout.addWidget(self.status_label)

        self.setLayout(main_layout)

    def select_all_options(self):
        self.confirm_cell_cb.setChecked(True)
        self.sheet_req_check_cb.setChecked(True)
        self.testcase_status_cb.setChecked(True)
        self.filename_check_cb.setChecked(True)
        self.sheet_check_cb.setChecked(True)
        self.check_contains_vietnamese_characters_cb.setChecked(True)
        self.check_invalid_text_cb.setChecked(True)
        self.check_incorrect_tb_content_cb.setChecked(True)

    def deselect_all_options(self):
        self.confirm_cell_cb.setChecked(False)
        self.sheet_req_check_cb.setChecked(False)
        self.testcase_status_cb.setChecked(False)
        self.filename_check_cb.setChecked(False)
        self.sheet_check_cb.setChecked(False)
        self.check_contains_vietnamese_characters_cb.setChecked(False)
        self.check_invalid_text_cb.setChecked(False)
        self.check_incorrect_tb_content_cb.setChecked(False)

    def on_folder_input_change(self, text):
        self.btn_execute.setEnabled(os.path.isdir(text.strip()))

    def select_folder(self):
        current_path = self.folder_input.text().strip()
        start_dir = (
            current_path if os.path.isdir(current_path) else os.path.expanduser("~")
        )

        if folder := QFileDialog.getExistingDirectory(self, "Select Folder", start_dir):
            self.folder_input.setText(folder)

    def start_execution(self):
        self.btn_export.setEnabled(False)
        folder_path = self.folder_input.text().strip()
        if not os.path.isdir(folder_path):
            QMessageBox.warning(
                self, "Invalid Folder", "Please provide a valid folder path."
            )
            return

        self.progress_bar.setValue(0)
        self.table.setRowCount(0)
        self.btn_execute.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.status_label.setText("Processing...")
        self.table.setSortingEnabled(False)

        options = {
            "check_invalid_sheets": self.sheet_check_cb.isChecked(),
            "check_filename_prefix": self.filename_check_cb.isChecked(),
            "check_required_sheets": self.sheet_req_check_cb.isChecked(),
            "check_confirm_cell": self.confirm_cell_cb.isChecked(),
            "check_testcase_status": self.testcase_status_cb.isChecked(),
            "check_contains_vietnamese_characters": self.check_contains_vietnamese_characters_cb.isChecked(),
            "check_invalid_text": self.check_invalid_text_cb.isChecked(),
            "check_incorrect_tb_content": self.check_incorrect_tb_content_cb.isChecked(),
        }

        load_config()
        self.worker = ExcelCheckWorker(folder_path, options)
        self.worker.progress_changed.connect(self.progress_bar.setValue)
        self.worker.file_result.connect(self.add_table_row)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.start()

    def stop_execution(self):
        if self.worker:
            self.btn_stop.setText("Stopping...")
            self.btn_stop.setEnabled(False)
            self.btn_execute.setEnabled(False)
            self.btn_export.setEnabled(False)
            self.progress_bar.setValue(0)
            self.status_label.setText("Process stopped by user... ")
            self.table.setSortingEnabled(False)
            QApplication.processEvents()
            self.worker.stop()
            self.btn_stop.setText("Stop")
            self.btn_execute.setEnabled(True)

    def add_table_row(self, prefix_path, path, status, error):
        row = self.table.rowCount()
        self.table.insertRow(row)

        items = [
            QTableWidgetItem(prefix_path.replace("/", "\\")),
            QTableWidgetItem(path),
            QTableWidgetItem(status),
            QTableWidgetItem(error),
        ]

        if status == "OK":
            items[2].setForeground(QColor("green"))
        elif status == "ERROR":
            items[2].setForeground(QColor("red"))

        for col, item in enumerate(items):
            self.table.setItem(row, col, item)

        if row == 0:
            self.btn_export.setEnabled(True)

    def open_selected_file(self, item):
        row = item.row()
        path = os.path.join(
            self.table.item(row, 0).text(), self.table.item(row, 1).text()
        )

        if os.path.exists(path):
            try:
                modifiers = QApplication.keyboardModifiers()
                if modifiers == Qt.ControlModifier:
                    folder_path = os.path.dirname(path)
                    if os.name == "nt":
                        os.startfile(folder_path)
                    elif sys.platform == "darwin":
                        subprocess.call(["open", folder_path])
                    else:
                        subprocess.call(["xdg-open", folder_path])
                else:
                    if os.name == "nt":
                        os.startfile(path)
                    elif sys.platform == "darwin":
                        subprocess.call(["open", path])
                    else:
                        subprocess.call(["xdg-open", path])
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Could not open: {str(e)}")
        else:
            QMessageBox.warning(self, "File Not Found", f"File not found: {path}")

    def on_finished(self):
        self.btn_execute.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.table.setSortingEnabled(True)
        self.table.sortItems(3, Qt.DescendingOrder)

        if self.worker._stop_event.is_set():
            self.status_label.setText("Process stopped by user")
            QMessageBox.information(self, "Stopped", "Process was stopped by user.")
        else:
            total_files = self.table.rowCount()
            error_count = sum(
                1
                for row in range(total_files)
                if self.table.item(row, 2).text() == "ERROR"
            )
            ok_count = total_files - error_count
            summary = f"Check completed.\nTotal files: {total_files}\nOK: {ok_count}\nErrors: {error_count}"
            self.status_label.setText("Process completed")
            QMessageBox.information(self, "Done", summary)

    def closeEvent(self, event):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
        event.accept()

    def export_results(self):
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "There are no results to export.")
            return

        default_name = (
            f"Excel_Check_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Results", default_name, "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Check Results"

            headers = ["Prefix Path", "Relative Path", "Status", "Errors"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for row in range(self.table.rowCount()):
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    ws.cell(
                        row=row + 2, column=col + 1, value=item.text() if item else ""
                    )

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            QMessageBox.information(
                self, "Success", f"Results exported to:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(
                self, "Export Error", f"Failed to export results:\n{str(e)}"
            )


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
