import sys
import os
import time
from openpyxl import load_workbook
from PyQt5.QtWidgets import (
    QApplication, QWidget, QPushButton, QVBoxLayout, QFileDialog,
    QLabel, QHBoxLayout, QProgressBar, QMessageBox, QLineEdit, QTableWidget, QTableWidgetItem
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QColor
import subprocess

# ----------------- Excel File Validation Logic -----------------
def check_excel_file_advanced(file_path):
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        filename = os.path.basename(file_path)
        results = []

        required_sheets = {'表紙', 'テスト項目'}
        missing_sheets = required_sheets - set(wb.sheetnames)
        if missing_sheets:
            return "ERROR", "Missing sheet(s): " + ", ".join(missing_sheets)

        ws = wb['表紙']
        p24_value = ws['P24'].value
        if p24_value is None or str(p24_value).strip() == "":
            wb.close()
            return "ERROR", "Missing Confirm by"

        ws2 = wb['テスト項目']
        error_rows = []
        checked_rows = []
        status_col = 0
        pref_col = 0

        for cell in ws2[3]:
            if cell.value == "確認":
                status_col = cell.column
            if cell.value == "参考":
                pref_col = cell.column

        max_rows_to_check = 1000
        consecutive_empty_limit = 10
        consecutive_empty = 0

        for row in range(5, max_rows_to_check + 1):
            if consecutive_empty >= consecutive_empty_limit:
                break
            b_cell = ws2.cell(row=row, column=2)
            b_value = b_cell.value
            if b_value is not None and str(b_value).strip():
                consecutive_empty = 0
                checked_rows.append(row)
                
                status_cell = ws2.cell(row=row, column=status_col)
                status_value = status_cell.value
                if status_value is None or str(status_value).strip().upper() != "OK":
                    error_rows.append(f"Row {row} (B{row}='{str(b_value).strip()}')")
                
                # pref_cell = ws2.cell(row=row, column=pref_col)
                # if pref_cell.hyperlink:
                    # link = pref_cell.hyperlink.target
                    # print(link, pref_col, pref_cell.value)
                    # if link.startswith("file:///"):
                    #     link_ref = link[8:]
                    #     if "!" in link_ref:
                    #         target_sheet = link_ref.split("!")[0].strip("'")
                    #         if target_sheet not in wb.sheetnames or target_sheet != pref_cell.value:
                    #             error_rows.append(f"Broken link in row {row}: {link_ref} not found in workbook or does not match cell value")
                    #     else:
                    #         error_rows.append(f"Broken link in row {row}: {link_ref} does not point to a valid sheet")
            else:
                consecutive_empty += 1

        if error_rows:
            wb.close()
            return "ERROR", f"{len(error_rows)} rows status != 'OK'"
        wb.close()
        return "OK", ""
    except Exception as e:
        return "ERROR", str(e)

def find_excel_files_recursive(folder_path):
    excel_files = []
    excel_extensions = ('.xlsx', '.xlsm', '.xls')
    for root_dir, _, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith(excel_extensions):
                excel_files.append(os.path.join(root_dir, file))
    return excel_files

# ----------------- Worker Thread -----------------
class ExcelCheckWorker(QThread):
    progress_changed = pyqtSignal(int)
    file_result = pyqtSignal(str, str, str)  # status, relative_path, error
    finished_signal = pyqtSignal()

    def __init__(self, folder_path):
        super().__init__()
        self.folder_path = folder_path

    def run(self):
        files = find_excel_files_recursive(self.folder_path)
        total = len(files)
        if not files:
            self.file_result.emit("INFO", "", "No Excel files found.")
            self.finished_signal.emit()
            return

        for i, file_path in enumerate(files, 1):
            relative_path = os.path.relpath(file_path, self.folder_path)
            status, error_msg = check_excel_file_advanced(file_path)
            self.file_result.emit(status, relative_path, error_msg)
            self.progress_changed.emit(int((i / total) * 100))
            time.sleep(0.05)

        self.finished_signal.emit()

# ----------------- PyQt Main Window -----------------
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Validator")
        self.setGeometry(100, 100, 800, 500)

        main_layout = QVBoxLayout()
        input_layout = QHBoxLayout()
        button_layout = QHBoxLayout()

        # Folder input bar
        self.folder_input = QLineEdit()
        self.folder_input.setPlaceholderText("Paste or type folder path here...")
        self.folder_input.textChanged.connect(self.on_folder_input_change)

        # Select button
        self.btn_select = QPushButton("Browse")
        self.btn_select.clicked.connect(self.select_folder)

        input_layout.addWidget(self.folder_input)
        input_layout.addWidget(self.btn_select)

        # Execute button
        self.btn_execute = QPushButton("Execute")
        self.btn_execute.setEnabled(False)
        self.btn_execute.clicked.connect(self.start_execution)

        button_layout.addWidget(self.btn_execute)

        # Table widget
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Status", "Path file", "Errors"])
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(self.table.SelectRows)
        self.table.itemDoubleClicked.connect(self.open_selected_file)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setAlignment(Qt.AlignCenter)
        self.progress_bar.setValue(0)

        # Layout add
        main_layout.addLayout(input_layout)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.table)
        main_layout.addWidget(self.progress_bar)
        self.setLayout(main_layout)

        self.worker = None

    def on_folder_input_change(self, text):
        self.btn_execute.setEnabled(os.path.isdir(text.strip()))

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            self.folder_input.setText(folder)

    def start_execution(self):
        folder_path = self.folder_input.text().strip()
        if not os.path.isdir(folder_path):
            QMessageBox.warning(self, "Invalid Folder", "Please provide a valid folder path.")
            return

        self.progress_bar.setValue(0)
        self.table.setRowCount(0)
        self.btn_execute.setEnabled(False)

        self.worker = ExcelCheckWorker(folder_path)
        self.worker.progress_changed.connect(self.progress_bar.setValue)
        self.worker.file_result.connect(self.add_table_row)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.start()

    def add_table_row(self, status, path, error):
        row = self.table.rowCount()
        self.table.insertRow(row)
        status_item = QTableWidgetItem(status)
        path_item = QTableWidgetItem(path)
        error_item = QTableWidgetItem(error)
        if status == "OK":
            status_item.setForeground(QColor("green"))
        elif status == "ERROR":
            status_item.setForeground(QColor("red"))
        self.table.setItem(row, 0, status_item)
        self.table.setItem(row, 1, path_item)
        self.table.setItem(row, 2, error_item)

    def open_selected_file(self, item):
        row = item.row()
        path = os.path.join(self.folder_input.text(), self.table.item(row, 1).text())
        if os.path.exists(path):
            if os.name == 'nt':
                os.startfile(path)
            elif sys.platform == 'darwin':
                subprocess.call(['open', path])
            else:
                subprocess.call(['xdg-open', path])
        else:
            QMessageBox.warning(self, "File Not Found", f"File not found: {path}")

    def on_finished(self):
        self.btn_execute.setEnabled(True)
        QMessageBox.information(self, "Done", "Validation completed.")

# ----------------- Entry Point -----------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
